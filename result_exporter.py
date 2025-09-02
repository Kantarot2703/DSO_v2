import pandas as pd
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def export_result_to_excel(results, output_path="output/result.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if isinstance(results, list):
        df = pd.DataFrame(results)
    elif isinstance(results, pd.DataFrame):
        df = results.copy()
    else:
        raise ValueError("❌ Invalid result format. Must be list or DataFrame.")

    def _norm_cell(v):
        if v is None:
            return "-"
        if isinstance(v, float) and pd.isna(v):
            return "-"
        if isinstance(v, str):
            s = v.strip()
            if s == "" or s in {"-", "–", "—"}:
                return "-"
        return v

    df = df.applymap(_norm_cell)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sheet_name = "Result"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="EDEDED")  
        thin = Side(style="thin", color="BFBFBF")            
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c in ws[1]:
            c.fill = header_fill
            c.font = Font(bold=True)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_r = ws.max_row
        max_c = ws.max_column
        for r in ws.iter_rows(min_row=2, max_row=max_r, min_col=1, max_col=max_c):
            for cell in r:
                cell.border = border
                if cell.value in (None, ""):
                    cell.value = "-"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, max_c + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

    print(f"✅ Exported result to: {output_path}")