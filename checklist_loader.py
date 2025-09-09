import os, io, re
import html as _html
import pandas as pd
import logging
import fitz, difflib
import unicodedata as _ud
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from collections import defaultdict


# Allowed part codes from PDF filenames
ALLOWED_PART_CODES = ['UU1_DOM', 'DOM', 'UU1', '2LB', '2XV', '4LB', '19L', '19A', '21A', 'DC1']

TOKEN_RE = re.compile(
    r"[A-Za-z0-9\u00C0-\u024F\u0400-\u04FF\u0E00-\u0E7F]+(?:-[A-Za-z0-9\u00C0-\u024F\u0400-\u04FF\u0E00-\u0E7F]+)?"
)

# ----- Font-size tolerance -----
def _size_meets_threshold(measured_mm: float, required_mm: float) -> bool:
    try:
        m = float(measured_mm or 0.0)
        r = float(required_mm or 0.0)
    except Exception:
        return False
    return (m + 1e-6) >= r

FONT_SOFT_UPPER_REL = 0.15 
FONT_SOFT_UPPER_ABS = 0.30 

def _oversize_warning(measured_mm: float, required_mm: float,
                      rel: float = FONT_SOFT_UPPER_REL, abs_mm: float = FONT_SOFT_UPPER_ABS):
    try:
        m = float(measured_mm or 0.0)
        r = float(required_mm or 0.0)
    except Exception:
        return (False, 0.0, 0.0)
    u_soft = r * (1.0 + rel) + abs_mm
    if m > u_soft:
        return (True, m - u_soft, u_soft)
    return (False, 0.0, u_soft)

def _contains_any(s: str, keys) -> bool:
    s = (s or "").lower()
    return any(k in s for k in keys)

def _pt_to_mm(pt: float) -> float:
    return (pt or 0.0) * 25.4 / 72.0

def _pick_size_mm(item: dict) -> float:
    if not item:
        return 0.0
    if "size_mm" in item and item.get("size_mm") is not None:
        try:
            return float(item["size_mm"])
        except Exception:
            return 0.0
    unit = str(item.get("size_unit") or "").lower()
    val = float(item.get("size", 0) or 0)
    return _pt_to_mm(val) if unit == "pt" else val

def _parse_threshold_to_mm(spec_text: str):
    if not isinstance(spec_text, str):
        return None
    s = spec_text.lower().strip()

    # ตัวอย่างที่รองรับ:
    m = re.search(r"(≥|<=|>=|<|>)\s*(\d+(?:\.\d+)?)\s*(mm\.?|pt\.?)?", s)
    if not m:
        return None
    val = float(m.group(2))
    unit = (m.group(3) or "mm").replace(".", "").strip()
    if unit.startswith("pt"):
        return (val * 25.4) / 72.0
    return val

def _fmt_mm(x: float) -> str:
    return f"{x:.1f} mm"

def _extract_underlined_substrings(html_text: str) -> list:
    if not isinstance(html_text, str) or not html_text.strip():
        return []
    parts = []
    for m in re.finditer(r"<u>(.*?)</u>", html_text, flags=re.IGNORECASE | re.DOTALL):
        frag = m.group(1) or ""
        frag_plain = re.sub(r"<[^>]+>", "", frag).strip()
        if frag_plain:
            parts.append(frag_plain)
    return parts

def _dedup_notes(notes):
    out, seen = [], set()
    for n in notes or []:
        s = str(n).strip()
        if not s or s == "-":
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out

def _dash_norm(x):
    s = "" if x is None else str(x).strip()
    if s == "" or s.lower() in {"-", "–", "—", "none", "nan"}:
        return "-"
    return s

def _is_all_caps_approx(s: str) -> bool:
    s = _ud.normalize("NFKC", str(s or ""))
    letters = [ch for ch in s if ch.isalpha()]
    if not letters:
        return False
    return all(ch == ch.upper() for ch in letters)

def _is_risky_term(term: str) -> bool:
    s = (term or "")
    s_nfkc = _ud.normalize("NFKC", s)
    alnum = [c for c in s_nfkc if c.isalnum()]
    has_non_alnum = any(not (c.isalnum() or c.isspace()) for c in s_nfkc)
    has_digit = any(c.isdigit() for c in s_nfkc)
    has_symbol = any(c in "+°®™×/%‐–—+-" for c in s_nfkc)
    return has_non_alnum or (len(alnum) <= 3) or (has_digit and has_symbol)

def _fuzzy_ratio(a: str, b: str) -> float:
    a_n = normalize_text(a)
    b_n = normalize_text(b)
    return difflib.SequenceMatcher(None, a_n, b_n).ratio()

def normalize_headers(df):
    rename = {}
    for col in df.columns:
        name = str(col).strip().lower()
        # Requirement
        if ("require" in name) or ("ข้อกำหนด" in name) or ("หัวข้อ" in name):
            rename[col] = "Requirement"
            continue
        # Term
        if ("symbol" in name) or ("exact" in name) or ("term" in name):
            rename[col] = "Symbol/ Exact wording"
            continue
        # Spec
        if ("spec" in name) or ("specification" in name):
            rename[col] = "Specification"
            continue

        # Package Panel
        if ("package panel" in name) or ("package" in name and "panel" in name):
            rename[col] = "Package Panel"
            continue

        # Procedure
        if ("procedure" in name) or ("process" in name):
            rename[col] = "Procedure"
            continue

        # Remark
        if ("remark" in name) or ("หมายเหตุ" in name):
            rename[col] = "Remark"
            continue

    df = df.rename(columns=rename)

    if "Requirement" not in df.columns:
        raise ValueError("Checklist Excel must contain a column recognizable as 'Requirement'.")
    return df

def get_strikeout_or_red_text_rows(excel_path, sheet_name, header_row_index):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    bad_rows = set()

    for row in ws.iter_rows(min_row=header_row_index + 2):  
        row_index = row[0].row
        for cell in row:
            font = cell.font
            color = font.color

            if not font:
                continue

            is_strike = font.strike
            is_red = False

            if color:
                if color.type == "rgb" and color.rgb:
                    is_red = color.rgb.upper().startswith("FF0000")
                elif color.type == "theme" and hasattr(color, "theme"):
                    if color.theme == 10:
                        is_red = True

            if is_strike and is_red:
                bad_rows.add(row_index)
                break 
    return bad_rows

def extract_part_code_from_pdf(pdf_filename):
    basename = os.path.basename(pdf_filename).upper().replace(" ", "").replace(",", "")
    found = []

    for code in ALLOWED_PART_CODES:
        if code in basename and code not in found:
            found.append(code)

    # สแกนเนื้อหา PDF หน้าแรกๆ เพิ่มเติม
    try:
        doc = fitz.open(pdf_filename)
        try:
            pages_to_scan = min(8, len(doc))
            tokens = set()
            for i in range(pages_to_scan):
                txt = (doc.load_page(i).get_text("text") or "").upper()
                for t in re.split(r"[^A-Z0-9_]+", txt):
                    t = t.strip()
                    if t:
                        tokens.add(t)

            # เติมโค้ดที่พบในเนื้อหา
            for code in ALLOWED_PART_CODES:
                if code in tokens and code not in found:
                    found.append(code)

            # ถ้ามีทั้ง UU1 และ DOM → ใส่ UU1_DOM ด้วย
            if ("UU1" in tokens and "DOM" in tokens) and ("UU1_DOM" not in found):
                found.insert(0, "UU1_DOM")
        finally:
            doc.close()
    except Exception:
        pass

    # คืนค่าแบบไม่ซ้ำ คงลำดับเดิม
    out = []
    for c in found:
        if c not in out:
            out.append(c)
    return out

def fuzzy_find_columns(df):
    term_col = None
    lang_col = None
    spec_col = None  

    logging.info(f"🧾 Columns found in sheet: {list(df.columns)}")

    # add Fallback mapping ถ้า fuzzy หาไม่เจอ
    def _norm(s: object) -> str:
         return str(s).replace("\xa0", " ").strip().lower()

    # ตรง Keyword ชัดเจน 
    for col in df.columns:
        if pd.isna(col): 
            continue
        n = _norm(col)
        n_compact = n.replace(" ", "")

        if term_col is None and any(k in n_compact for k in ["term", "ข้อความ", "exactwording", "symbol", "wording"]):
            term_col = col

        if lang_col is None and any(k in n_compact for k in ["languagecode", "langcode", "language", "lang", "ภาษา"]):
            lang_col = col

        if spec_col is None and any(k in n_compact for k in ["specification", "spec", "requirement", "ข้อกำหนด"]):
            spec_col = col

    # Fallback ผ่อนเงื่อไข
    if term_col is None:
        for col in df.columns:
            n = _norm(col)
            n_compact = n.replace(" ", "")
            if any(k in n_compact for k in ["term", "symbol", "exact", "wording"]):
                term_col = col
                break

    if lang_col is None:
        for col in df.columns:
            n = _norm(col)
            if ("language" in n) or (n == "lang") or ("language code" in n) or ("lang code" in n):
                lang_col = col
                break

    if spec_col is None:
        for col in df.columns:
            n = _norm(col)
            if (n == "spec") or ("specification" in n) or ("requirement" in n) or ("ข้อกำหนด" in n):
                spec_col = col
                break

    # OPTIONAL บังคับตั้งชื่อคอลัมน์ Requirement ถ้าตรวจพบ
    for c in df.columns:
        if str(c).strip().lower() == "requirement":
            if c != "Requirement":
                df = df.rename(columns={c: "Requirement"})
            break

    logging.info(f"🛟 Fallback columns → Term: {term_col}, Language: {lang_col}, Spec: {spec_col}")
    return term_col, lang_col, spec_col

_FULL2HALF = str.maketrans({
    "＋": "+", "﹢": "+", "⁺": "+", "₊": "+", "➕": "+", 
    "－": "-", "％": "%", "＝": "=", "＊": "*", "／": "/",
    "＇": "'", "＂": '"', "＆": "&", "｜": "|", "＃": "#", "＠": "@",
    "（": "(", "）": ")", "［": "[", "］": "]", "｛": "{", "｝": "}",
    "，": ",", "．": ".", "：": ":", "；": ";", "！": "!", "？": "?",
    "～": "~", "＾": "^", "｀": "`", "＿": "_", "＜": "<", "＞": ">",
    "　": " ", 
})

def normalize_text(text: str) -> str:
    if text is None:
        return ""
    s = str(text)
    s = _ud.normalize("NFKD", s)
    s = "".join(ch for ch in s if not _ud.combining(ch))
    s = s.translate(_FULL2HALF)
    s = s.replace("\u00A0", " ")
    s = s.replace("‐", "-").replace("–", "-").replace("—", "-")
    s = s.lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s

# --- Country equivalents (normalize แล้ว) ---
_TH_EQ = {
    "thailand","thailande","tailandia","tailândia","thailandia","tajlandia",
    "thajsko","thaimaa","tayland","тайланд","таиланд","ταϊλάνδη","تايلاند"
}
def _extract_th_country_flag(term_text: str) -> bool:
    return any(k in normalize_text(term_text) for k in _TH_EQ)

def _must_contain_country_th(text_norm: str) -> bool:
    return any(k in text_norm for k in _TH_EQ)

def extract_underlines_from_excel(excel_path, sheet_name, header_row_index, df):
    try:
        term_col_name = None
        for cand in ["Symbol/Exact wording", "Symbol/ Exact wording"]:
            if cand in df.columns:
                term_col_name = cand
                break
        if not term_col_name:
            df["__Term_HTML__"] = df.get("Symbol/Exact wording", df.get("Symbol/ Exact wording", "")).astype(str)
            return df

        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=False)
        ws = wb[sheet_name]

        # แถวข้อมูลแรกในชีตโดยประมาณ:
        approx_start_row = header_row_index + 2 
        term_col_idx = list(df.columns).index(term_col_name) + 1

        def _cell_rich_to_html(cell, plain_text: str) -> str:

            txt_plain = _html.escape(plain_text or "")

            try:
                val = getattr(cell, "value", None)

                rt = getattr(val, "richText", None)
                if rt:
                    parts = []
                    for r in rt:
                        t = getattr(r, "text", "")
                        f = getattr(r, "font", None)
                        is_bold = bool(getattr(f, "bold", False)) if f else False
                        is_ul   = bool(getattr(f, "underline", False)) if f else False
                        h = _html.escape(str(t))
                        if is_bold: h = f"<b>{h}</b>"
                        if is_ul:   h = f"<u>{h}</u>"
                        parts.append(h)
                    return "".join(parts)

                runs = []
                if hasattr(val, "rich") and val.rich:
                    runs = val.rich
                elif hasattr(val, "runs") and val.runs:
                    runs = val.runs

                if runs:
                    parts = []
                    for r in runs:
                        t = getattr(r, "text", None)
                        if t is None:
                            t = str(r)
                        h = _html.escape(str(t))
                        f = getattr(r, "font", None)
                        is_bold = bool(getattr(f, "bold", False)) if f else False
                        is_ul   = bool(getattr(f, "underline", False)) if f else False
                        if is_bold: h = f"<b>{h}</b>"
                        if is_ul:   h = f"<u>{h}</u>"
                        parts.append(h)
                    return "".join(parts)

            except Exception as e:
                logging.debug(f"[underline rich] parse failed: {e}")

            try:
                f = getattr(cell, "font", None)
                if f:
                    if bool(getattr(f, "bold", False)):
                        txt_plain = f"<b>{txt_plain}</b>"
                    if bool(getattr(f, "underline", False)):
                        txt_plain = f"<u>{txt_plain}</u>"
            except Exception:
                pass
            return txt_plain

        def _norm(s):
            t = "" if s is None else str(s)
            t = t.replace("\r", " ").replace("\n", " ")
            t = " ".join(t.split())
            return t.strip()

        probe_window = 6
        delta_candidates = []
        for i in range(min(len(df), probe_window)):
            df_txt = _norm(df.iloc[i].get(term_col_name, ""))
            if not df_txt:
                continue

            center = approx_start_row + i
            found_delta = None
            for d in range(-5, 6):
                r = center + d
                if r < 1:
                    continue
                cell_txt = _norm(ws.cell(row=r, column=term_col_idx).value)
                if cell_txt == df_txt:
                    found_delta = d
                    break
            if found_delta is not None:
                delta_candidates.append(found_delta)

        real_start_row = approx_start_row
        if delta_candidates:
            delta_candidates.sort()
            real_delta = delta_candidates[len(delta_candidates)//2]
            real_start_row = approx_start_row + real_delta

        logging.info(f"[underline] approx_start_row={approx_start_row}, real_start_row={real_start_row}")

        # สร้าง HTML ตามแถวจริง 
        html_list = []
        for i in range(len(df)):
            ws_row = real_start_row + i
            cell = ws.cell(row=ws_row, column=term_col_idx)

            # ใช้ข้อความ "ดิบ" จาก df ไม่ normalize ทิ้งบรรทัดใหม่
            text_raw = df.iloc[i].get(term_col_name, "")
            if text_raw is None:
                text_raw = ""
            text_raw = str(text_raw)

            html_list.append(_cell_rich_to_html(cell, text_raw))

        df["__Term_HTML__"] = [(h or "").replace("\r", "").replace("\n", "<br>") for h in html_list]
        return df

    except Exception as e:
        logging.debug(f"Underline extraction failed: {e}")
        term_col_name = "Symbol/Exact wording" if "Symbol/Exact wording" in df.columns else "Symbol/ Exact wording"
        df["__Term_HTML__"] = df.get(term_col_name, "").astype(str)
        return df

    except Exception as e:
        logging.debug(f"Underline extraction failed: {e}")
        df["__Term_HTML__"] = df.get("Symbol/Exact wording", "").astype(str)
        return df

def load_checklist(excel_path, pdf_filename=None):
    all_sheets = pd.read_excel(excel_path, sheet_name=None)
    sheet_names = list(all_sheets.keys())

    if not pdf_filename:
        raise ValueError("📄 กรุณาอัปโหลดไฟล์ PDF ก่อน เพื่อจับคู่กับ Sheet ของ Checklist")

    part_codes = extract_part_code_from_pdf(pdf_filename)
    if not part_codes:
        raise ValueError("❌ ไม่พบ Part code ที่ระบุไว้ในชื่อไฟล์ PDF")

    logging.info(f"📂 PDF filename: {pdf_filename}")
    logging.info(f"🧠 Part codes detected: {part_codes}")
    logging.info(f"📄 Sheet names: {sheet_names}")

    for sheet_name in sheet_names:
        sheet_name_normalized = sheet_name.upper().replace(" ", "")
        for code in part_codes:
            if sheet_name_normalized.startswith(code):
                logging.info(f"✅ Found matching sheet: {sheet_name}")
                df = all_sheets[sheet_name]

                HEADER_HINTS = [
                    r"\brequirement\b",
                    r"\blanguage\b|\blang\b|\blanguage\s*code\b",
                    r"\bsymbol\b|\bexact\s*wording\b|\bterm\b",
                    r"\bspec(ification)?\b",
                ]

                def _row_score(cells):
                    score = 0
                    for c in cells:
                        s = "" if c is None else str(c).strip()
                        s_norm = s.lower()
                        if not s:
                            continue

                        # ให้แต้มถ้าตรงคีย์เวิร์ดหัวตาราง
                        if any(re.search(p, s_norm) for p in HEADER_HINTS):
                            score += 5
                        if len(s) <= 24:
                            score += 1
                        if "=" in s or "“" in s or "”" in s:
                            score -= 2
                    return score
                
                # ค้นหา header ภายใน 15 แถวแรก เลือกแถวที่ได้คะแนนมากสุด
                header_row_index = None
                best_score = -10
                scan_upto = min(15, len(df))
                for i in range(scan_upto):
                    row_vals = list(df.iloc[i].values)
                    if pd.Series(row_vals).notna().sum() < 2:
                        continue
                    sc = _row_score(row_vals)
                    if sc > best_score:
                        best_score = sc
                        header_row_index = i

                if header_row_index is None:
                    raise ValueError(f"❌ ไม่พบแถว header ที่เหมาะสมใน sheet: {sheet_name}")
                
                df.columns = df.iloc[header_row_index]
                df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
                df = df[header_row_index + 1:].reset_index(drop=True)
                logging.info(f"🧾 Header chosen at row {header_row_index+1} | columns: {list(df.columns)[:6]}...")

                # Normalize header names (เพิ่ม mapping ให้คอลัมน์ Verification)
                _ren = {}
                for c in list(df.columns):
                    n = str(c).strip().lower()
                    if "verify" in n or "verification" in n or "ตรวจ" in n:  # ครอบคลุม EN/TH
                        _ren[c] = "Verification"
                if _ren:
                    df = df.rename(columns=_ren)

                # Column Mapping
                term_col, lang_col, spec_col = fuzzy_find_columns(df)
                logging.info(f"🔎 ใช้คอลัมน์ Term: {term_col}, Language: {lang_col}, Spec: {spec_col}")

                # Standardize
                if spec_col and spec_col in df.columns:
                    df[spec_col] = df[spec_col].apply(
                        lambda x: "-" if pd.isna(x) or str(x).strip().upper() in ["N/A", "NONE", "-"] else str(x)
                    )

                # ffill เฉพาะคอลัมน์ที่มีอยู่จริง
                columns_to_ffill = [c for c in df.columns if str(c).strip().lower() in ["requirement", "language"]]
                if columns_to_ffill:
                    df[columns_to_ffill] = df[columns_to_ffill].ffill()

                # Rename ถ้า term_col ไม่อยู่ ให้สร้างคอลัมน์ว่างกัน KeyError
                if term_col in df.columns:
                    df = df.rename(columns={term_col: "Symbol/Exact wording"})
                elif "Symbol/Exact wording" not in df.columns:
                    df["Symbol/Exact wording"] = "-"

                GROUP_RE = re.compile(r"^\s*\[GROUP:\s*(?P<name>.+?)\s*\]\s*\[(?P<mode>ANY|ALL)\]\s*$", re.IGNORECASE)

                def _split_simple_list(cell: str):
                    """รองรับหลายพาธคั่นด้วย ; | หรือขึ้นบรรทัดใหม่"""
                    if not isinstance(cell, str):
                        return []
                    s = cell.strip()
                    if not s or s in ["-", "N/A", "None"]:
                        return []
                    parts = re.split(r"[;\n|]", s.replace("\r", ""))
                    return [p.strip().replace("\\", "/") for p in parts if p.strip()]
                
                def _parse_image_groups(cell: str):
                    """
                    รูปแบบที่รองรับ:
                    - แบบมี group/tag:
                        [GROUP: Old logo][ALL]
                        //server/share/old1.png
                        //server/share/old2.png
                        [GROUP: New logo][ANY]
                        assets/new1.png
                        assets/new2.png
                    - แบบธรรมดา: หลายพาธในเซลล์เดียว -> กลุ่มเดียว mode=ANY
                    - สมมติว่ามีคอลัมน์ Image Match แยก: จะไป normalize ต่อ
                    """
                    if not isinstance(cell, str) or not cell.strip():
                        return []
                    lines = [ln.strip() for ln in cell.replace("\r", "").split("\n")]
                    groups, cur = [], None
                    for ln in lines:
                        if not ln:
                            continue
                        m = GROUP_RE.match(ln)
                        if m:
                            if cur and cur["paths"]:
                                groups.append(cur)
                            cur = {"name": m.group("name"), "mode": m.group("mode").lower(), "paths": []}
                        else:
                            for p in re.split(r"[;|]", ln):
                                p = p.strip()
                                if p and p not in ["-", "N/A", "None"]:
                                    if cur is None:
                                        cur = {"name": "", "mode": "any", "paths": []}
                                    cur["paths"].append(p.replace("\\", "/"))
                    if cur and cur["paths"]:
                        groups.append(cur)

                    # ถ้าไม่มี [GROUP:...] เลย และมีพาธเดียว/หลายพาธ -> กลุ่มเดียว ANY 
                    if not groups:
                        paths = _split_simple_list(cell)
                        return [{"name": "", "mode": "any", "paths": paths}] if paths else []
                    return groups
                
                def _flatten_paths(groups):
                    out = []
                    for g in (groups or []):
                        out.extend(g.get("paths", []))
                    return out 
                
                # สร้างคอลัมน์ Image_Groups + Image_Paths_Flat
                if "Image Path" in df.columns:
                    df["Image_Groups"] = df["Image Path"].fillna("").astype(str).apply(_parse_image_groups)
                    df["Image_Paths_Flat"] = df["Image_Groups"].apply(_flatten_paths)
                else:
                    df["Image_Groups"] = [[] for _ in range(len(df))]
                    df["Image_Paths_Flat"] = [[] for _ in range(len(df))]

                # หากมีคอลัมน์ Image Match แยก (ANY/ALL) ให้บังคับโหมดกลุ่มเดี่ยวให้ตรงค่านี้
                if "Image Match" in df.columns:
                    def _apply_global_mode(groups, mode_cell):
                        mode = (str(mode_cell).strip().lower() if isinstance(mode_cell, str) else "")
                        if mode in ["all", "any"]:
                            # ถ้ามีหลายกลุ่ม จะ apply ให้ทุกกลุ่ม
                            for g in groups or []:
                                g["mode"] = mode
                        return groups
                    df["Image_Groups"] = [
                        _apply_global_mode(g, m) for g, m in zip(df["Image_Groups"], df["Image Match"])
                    ] 

                # Add extract hyperlink targets from Remark cells
                try:
                    wb = load_workbook(excel_path, data_only=True)
                    ws = wb[sheet_name]

                    if "Remark" in df.columns:
                        remark_col_idx = list(df.columns).index("Remark") + 1
                        start_row = header_row_index + 2 

                        remark_links = []
                        for i in range(len(df)):
                            cell = ws.cell(row=start_row + i, column=remark_col_idx)
                            url = ""
                            if cell and cell.hyperlink:
                                url = cell.hyperlink.target or ""
                            remark_links.append(url)
                        df["Remark Link"] = remark_links
                    else:
                        df["Remark Link"] = ""
                except Exception as _e:
                    logging.debug(f"Remark hyperlink extraction skipped: {_e}")
                    df["Remark Link"] = ""

                # RESOLVE absolute paths for Image_Groups and add _HasImage BEFORE filtering
                excel_dir = os.path.dirname(excel_path)

                def _resolve_group_paths(groups):
                    out = []
                    for g in (groups or []):
                        paths = []
                        for p in g.get("paths", []):
                            if not isinstance(p, str) or not p.strip():
                                continue
                            p2 = p.strip().replace("\\", os.sep).replace("/", os.sep)
                            if not os.path.isabs(p2):
                                p2 = os.path.abspath(os.path.join(excel_dir, p2))
                            paths.append(p2)
                        out.append({"name": g.get("name",""), "mode": (g.get("mode") or "any").lower(), "paths": paths})
                    return out

                df["Image_Groups_Resolved"] = df["Image_Groups"].apply(_resolve_group_paths)

                def _has_any_image(groups):
                    if not groups:
                        return False
                    for g in groups:
                        for p in g.get("paths", []):
                            if isinstance(p, str) and p.strip():
                                return True
                    return False

                df["_HasImage"] = df["Image_Groups_Resolved"].apply(_has_any_image)

                # ถ้ามีคอลัมน์ Image Path แบบเดี่ยว ก็ resolve ให้ด้วย
                if "Image Path" in df.columns:
                    df["Image Path"] = df["Image Path"].fillna("").astype(str)
                else:
                    df["Image Path"] = ""

                def _resolve_path(p):
                    if not isinstance(p, str) or not p.strip():
                        return ""
                    p = p.strip().replace("\\", os.sep).replace("/", os.sep)
                    if os.path.isabs(p):
                        return p
                    return os.path.abspath(os.path.join(excel_dir, p))

                df["Image Path Resolved"] = df["Image Path"].apply(_resolve_path)
                
                # ตัดแถวว่าง/แถวผีหลัง
                def _clean(s):
                    return str(s).strip().lower()

                # ชื่อคอลัมน์สำคัญ (บางไฟล์อาจไม่มี spec_col)
                term_col_safe = "Symbol/Exact wording"
                spec_col_safe = spec_col if spec_col in df.columns else None

                # เงื่อนไขว่าง
                term_empty  = df[term_col_safe].astype(str).str.strip().isin(["", "-", "nan", "none", "n/a"])
                if spec_col_safe:
                    spec_empty  = df[spec_col_safe].astype(str).str.strip().isin(["", "-", "nan", "none", "n/a"])
                else:
                    spec_empty  = pd.Series([True] * len(df), index=df.index) # Series ของ True ยาวเท่า df (ให้ผ่านเงื่อนไขนี้ไป)

                # สร้าง Series ว่างสำหรับ Remark ถ้าไม่มีคอลัมน์
                remark_series = df["Remark"] if "Remark" in df.columns else pd.Series([""] * len(df), index=df.index)
                remark_empty = remark_series.astype(str).str.strip().isin(["", "-", "nan"])

                # Add ลิงก์ Remark ก็ถือว่า "มีข้อมูล"
                remark_link_series = df.get("Remark Link", pd.Series([""] * len(df), index=df.index))
                remark_link_empty = remark_link_series.astype(str).str.strip().isin(["", "-", "nan", ""])

                # Force keep row
                force_keep_mask = df.get("Requirement", pd.Series([""]*len(df))).astype(str).str.strip().str.lower() \
                                    .str.contains(r"instruction\s+of\s+play\s+function\s+feature", regex=True)
                
                # แถวที่ควรเก็บ = อย่างน้อยต้องมี Term หรือ Spec หรือ Remark ไม่ว่าง
                keep_mask = ~(term_empty & spec_empty & remark_empty & remark_link_empty) | df["_HasImage"] | force_keep_mask
                df = df[keep_mask].reset_index(drop=True)

                # กันช่องว่างล้วน (ยกเว้น Requirement/Language) เป็น NaN ล้วน
                non_struct_cols = [c for c in df.columns if str(c).strip().lower() not in ["requirement", "language"]]
                df = df[~df[non_struct_cols].isna().all(axis=1)].reset_index(drop=True)

                # Drop Red+Strike Rows (คง log ไว้ แต่ให้เป็น debug)
                bad_row_numbers = get_strikeout_or_red_text_rows(excel_path, sheet_name, header_row_index)
                logging.debug(f"❌ Red+Strike rows from Excel: {bad_row_numbers}")

                df["ExcelRow"] = df.index + header_row_index + 2 
                df.drop(columns=["ExcelRow"], inplace=True)

                df = extract_underlines_from_excel(excel_path, sheet_name, header_row_index, df)
                
                # อย่า explode อีกต่อไป — เก็บเป็น “สตริงเดียว” (จะมี \n ก็ปล่อยให้เป็นบรรทัดใหม่ในสตริง)
                df["Symbol/Exact wording"] = df["Symbol/Exact wording"].astype(str).str.replace("\r", "", regex=False)

                if "__Term_HTML__" in df.columns:
                    df["__Term_HTML__"] = df["__Term_HTML__"].apply(lambda x: str(x))

                # Language List 
                if lang_col:
                    df = df.rename(columns={lang_col: "Language"})
                    df["Language List"] = df["Language"].apply(lambda x: str(x).split(",") if pd.notna(x) else [])
                else:
                    df["Language List"] = [[] for _ in range(len(df))]

                # Extract from Remark 
                if "Remark" in df.columns:
                    def extract_languages_from_remark(remark, term):
                        langs = []
                        if pd.isna(remark): return langs
                        for line in str(remark).splitlines():
                            if "=" in line:
                                left, right = line.split("=", 1)
                                if term.strip().lower() in left.strip().lower():
                                    langs.append(right.strip())
                        return langs
                    df["Language List"] = [
                        extract_languages_from_remark(remark, term) or ["Unspecified"]
                        for remark, term in zip(df.get("Remark", []), df["Symbol/Exact wording"])
                    ]

                return df 

    raise ValueError("❌ ไม่พบ Sheet ที่ตรงกับ Part code จากชื่อไฟล์ PDF")

def start_check(df_checklist, extracted_text_list):

    logger = logging.getLogger(__name__)

    results = []
    grouped = defaultdict(list)

    skip_keywords = [
        "do not print", "do not forget", "see template", "don't forget",
        "for reference", "note", "click", "reminder", "refer template", "remove from template", "remove from mb legal template"
    ]

    manual_keywords = [
        #EN
        "brand logo", "copyright for t&f", "space for date code", "lion mark", "lionmark", "lion-mark", "ce mark", "en 71",
        "ukca", "mc mark", "cib", "argentina logo", "brazilian logo", "italy requirement", "france requirement",
        "sorting & donation label", "spain sorting symbols", "usa warning statement", "generic name",
        "upc requirement", "list of content : text", "list of content : pictorial", "product’s common", "product's common",
        # TH
        "โลโก้", "ลิขสิทธิ์", "ตรา", "สัญลักษณ์", "เครื่องหมาย",
        "สำรองพื้นที่รหัสวันที่", "เครื่องหมายรับรอง"
    ]

    # ใช้ Part No. เป็นเกณฑ์หลักในการคัดหน้า artwork 
    PARTNO_RE = re.compile(r"\b[A-Z]{3}\d{2}\b")

    doc_has_any_partno = False
    for _page in extracted_text_list:
        _txt = " ".join((it.get("text") or "") for it in _page).upper()
        if PARTNO_RE.search(_txt):
            doc_has_any_partno = True
            break

    # คัดหน้า artwork เท่านั้น (ตัดหน้าปก/หน้าไม่ใช่งานศิลป์ออก)
    def is_artwork_page(page_items):
        if not page_items:
            return False

        page_text = " ".join((it.get("text") or "") for it in page_items)
        page_up = page_text.upper()

        if PARTNO_RE.search(page_up):
            return True

        pdf_items = [it for it in page_items if (it.get("source") or "pdf").lower() != "ocr"]
        ocr_items = [it for it in page_items if (it.get("source") or "").lower() == "ocr"]

        def _has_big(items):
            try:
                return any((_pick_size_mm(it) >= 1.6) for it in items if isinstance(it, dict))
            except Exception:
                return False

        many_pdf   = len(pdf_items) >= 12
        many_words = len(page_text.split()) >= 20
        if (many_pdf and _has_big(pdf_items)) or (len(ocr_items) >= 12 and (_has_big(ocr_items) or _has_big(page_items))):
            return True

        MADE_IN_HINTS = (
            "made in","hecho en","fabriqué en","fabrique en","prodotto in","fabbricato in",
            "hergestellt in","gemaakt in","tillverkad i","valmistettu","fremstillet i","produceret i",
            "produsert i","wyprodukowano w","vyrobeno v","vyrobené v","gyártva","készült",
            "сделано в","произведено в","κατασκευ","παράγ","üret","صنع في","صناعة"
        )
        pn = normalize_text(page_text)
        return any(h in pn for h in MADE_IN_HINTS)

    artwork_pages = []
    page_mapping  = {}

    for real_idx, page_items in enumerate(extracted_text_list):
        consider = is_artwork_page(page_items)
        if not consider and doc_has_any_partno and real_idx > 0:
            consider = True
        if not consider and (not doc_has_any_partno) and real_idx > 0:
            consider = True
        if consider:
            artwork_pages.append(page_items)
            page_mapping[len(artwork_pages)] = real_idx + 1

    logger.info(
        "📄 Pages considered: %d (real pages: %s)",
        len(artwork_pages),
        list(page_mapping.values())
    )

    logger.info(
    "📄 Artwork-like pages considered: %d (real pages: %s)",
    len(artwork_pages),
    list(page_mapping.values())
    )

    # รวมข้อความแบบ normalize ทั้งหน้า เพื่อจับกรณีประโยคยาวข้ามบรรทัด
    page_norm_text = {}
    for artwork_index, page_items in enumerate(artwork_pages):
        real_no = page_mapping[artwork_index + 1]
        page_norm_text[real_no] = " ".join(
            normalize_text(it.get("text", "")) for it in page_items if (it.get("text"))
        )

    all_texts = []
    for artwork_index, page_items in enumerate(artwork_pages):
        for item in page_items:
            text_norm = normalize_text(item.get("text", ""))
            page_number = page_mapping[artwork_index + 1]
            all_texts.append((text_norm, page_number, item))

    # ---------- Language CODE from PDF CONTENT ----------
    def _tokens_upper(s: str) -> set:
        return {t for t in re.split(r"[^A-Z0-9_]+", (s or "").upper()) if t}

    def _detect_lang_codes_in_text(text: str) -> set:
        tokens = _tokens_upper(text)
        codes = set()

        for code in ALLOWED_PART_CODES:
            if code in tokens:
                codes.add(code)

        for m in re.findall(r"\b([2-9]LB)\b", (text or "").upper()):
            codes.add(m)
        return codes

    def _compact_pages(nums) -> str:
        nums = sorted(set(int(n) for n in nums))
        if not nums:
            return "-"
        out = []
        start = prev = nums[0]
        for n in nums[1:]:
            if n == prev + 1:
                prev = n
            else:
                out.append(str(start) if start == prev else f"{start}-{prev}")
                start = prev = n
        out.append(str(start) if start == prev else f"{start}-{prev}")
        return ", ".join(out)

    # Map: ภาพรวม code ต่อ "เลขหน้าใน PDF จริง" (เฉพาะหน้า artwork)
    page_langcodes = {}                 
    code_pages_map = defaultdict(set)

    for artwork_index, page_items in enumerate(artwork_pages):
        real_no = page_mapping[artwork_index + 1]      # เลขหน้าใน PDF จริง
        whole_text = " ".join([(it.get("text") or "") for it in page_items])
        det_codes = _detect_lang_codes_in_text(whole_text)
        page_langcodes[real_no] = det_codes
        for c in det_codes:
            code_pages_map[c].add(real_no)

    # แปลงรายชื่อหน้า All Pages 
    def _format_pages_for_output(found_pages):
        artwork_set = set(page_mapping.values())
        p = sorted(set(int(x) for x in (found_pages or [])) & artwork_set)
        if not artwork_set:
            return ", ".join(str(x) for x in p) if p else "-"

        coverage = (len(p) / max(1, len(artwork_set)))
        if coverage >= 0.90:
            return "All Pages"
        return ", ".join(str(x) for x in p) if p else "-"

    # --- helper: log หลักฐานการตรวจ สำหรับ term หนึ่งแถว ---
    def _log_evidence(term_text, spec_text, found_pages_list, matched_items, thr_mm, max_size_mm, page_str):
        pdf_hits = sum(1 for i in matched_items if (i.get("source") or "pdf").lower() != "ocr")
        ocr_hits = sum(1 for i in matched_items if (i.get("source") or "").lower() == "ocr")
        ul_hits  = sum(1 for i in matched_items if bool(i.get("underline")))
        bd_hits  = sum(1 for i in matched_items if bool(i.get("bold")))

        logger.info("🔎 [VERIFY] Term=%r | Spec=%r", term_text, spec_text or "-")
        logger.info("    ↳ Pages(evidence)=%s | Items=%d (PDF=%d, OCR=%d, Underline=%d, Bold=%d)",
                    page_str, len(matched_items), pdf_hits, ocr_hits, ul_hits, bd_hits)
        
        if thr_mm is not None:
            if max_size_mm is None:
                logger.info("    [FONT] Spec ≥ %.2f mm | observed: - (no measurable item)", float(thr_mm))
            else:
                ok = (max_size_mm >= thr_mm)
                logger.info("    [FONT] Spec ≥ %.2f mm | observed: %.2f mm → %s",
                            float(thr_mm), float(max_size_mm), "PASS" if ok else "FAIL")

        for i, it in enumerate(matched_items[:3]):
            logger.info("    [HIT#%d] text=%r | src=%s | underline=%s | bold=%s | size=%.2f mm",
                        i+1,
                        (it.get('text') or "")[:80],
                        (it.get('source') or "pdf"),
                        bool(it.get('underline')), bool(it.get('bold')),
                        _pick_size_mm(it))

    for idx, row in df_checklist.iterrows():
        requirement = str(row.get("Requirement", "")).strip()
        spec = str(row.get("Specification", "")).strip()
        package_panel = (str(row.get("Package Panel", "")) or "").strip() or "-"
        procedure = (str(row.get("Procedure", "")) or "").strip() or "-"
        remark_text = (str(row.get("Remark", "")) or "").strip()
        remark_link = (str(row.get("Remark Link", "")) or "").strip()

        # Normalize
        req_norm = normalize_text(requirement)
        spec = "-" if spec.lower() in ["", "n/a", "none", "unspecified", "nan"] else spec
        spec_norm = normalize_text(spec)

        spec_lower = spec.lower() if isinstance(spec, str) else ""

        # Term (ไม่ดึงค่าจากแถวอื่น)
        term_raw = row.get("Symbol/Exact wording", None)
        term_cell_raw = str(term_raw) if pd.notna(term_raw) else ""
        term_cell_clean = term_cell_raw.strip()
        term_cell_clean = "-" if term_cell_clean.lower() in ["", "n/a", "none", "unspecified", "nan"] else term_cell_clean
        
        term_html = str(row.get("__Term_HTML__", "") or "")
        term_lines = []
        if term_cell_clean != "-":
            term_lines.append(term_cell_clean)
            if term_html.strip():
                try:
                    under_parts = _extract_underlined_substrings(term_html)  
                    for up in under_parts:
                        if up and up not in term_lines:
                            term_lines.append(up)
                except Exception:
                    pass
        else:
            term_lines = []

        # HARD SKIP: ถ้าแถวถูกระบุ Manual ใน Excel ให้ข้ามการตรวจทั้งหมด
        raw_verif = (str(row.get("Verification", "")) or "").strip().lower()
        if raw_verif == "manual":
            grouped[(requirement, spec, "Manual")].append({
                "Term": term_cell_raw,
                "Found": "-",
                "Match": "-",
                "Pages": "-",
                "Font Size": "-",
                "Note": "-",
                "Verification": "Manual",
                "Remark": remark_text,
                "Remark URL": remark_link,
                "Package Panel": package_panel,
                "Procedure": procedure,
                "__Term_HTML__": row.get("__Term_HTML__", ""),
                "Image_Groups_Resolved": row.get("Image_Groups_Resolved", row.get("Image_Groups", [])),
            })
            continue

        # ข้าม row ไม่จำเป็น
        if any(kw in req_norm for kw in skip_keywords) or any(kw in spec_norm for kw in skip_keywords):
            continue

        remark_norm = normalize_text(remark_text)
        term_norm   = normalize_text(term_cell_clean)
        fields_norm = " ".join([req_norm, spec_norm, remark_norm])

        is_manual = any(kw in fields_norm for kw in manual_keywords)

        # Force manual เมื่อไม่มี term แต่มีภาพ (ไว้รอ OCR ภายหลัง)
        if not term_lines:
            is_manual = True

        # ถ้ามีภาพ และเจอคำที่ส่อว่าเป็นโลโก้/ลิขสิทธิ์ → Manual
        if bool(row.get("_HasImage", False)) and any(
            k in fields_norm for k in ["logo", "mark", "symbol", "copyright", "t&f", "t & f", "©", "™", "®"]
        ):
            is_manual = True

        # MANUAL SECTION 
        if is_manual:
            remark_str = str(row.get("Remark", "")).strip().lower()
            is_term_empty   = (len(term_lines) == 0)
            is_spec_empty   = (str(spec)).strip().lower() in ["", "-", "nan", "none", "null"]
            is_remark_empty = (remark_str in ["", "-", "nan"])

            if is_term_empty and is_spec_empty and is_remark_empty:
                if bool(row.get("_HasImage", False)):
                    grouped[(requirement, spec, "Manual")].append({
                        "Term": "-",
                        "Remark": remark_text or "-",
                        "Remark URL": remark_link or "-",
                        "Found": "-",
                        "Match": "-",
                        "Pages": "-",
                        "Font Size": "-",
                        "Note": "Manual check required",
                        "Verification": "Manual",
                        "Package Panel": package_panel,
                        "Procedure": procedure,
                        "__Term_HTML__": row.get("__Term_HTML__", ""),
                        "Image_Groups_Resolved": row.get("Image_Groups_Resolved", row.get("Image_Groups", [])),
                    })
                    continue
                else:
                    continue

            logger.info(
                "🟨 [MANUAL] Req: '%s' | Spec: '%s' → Manual verification",
            requirement, (spec or "-")
            )

            if not term_lines:
                grouped[(requirement, spec, "Manual")].append({
                    "Term": term_cell_raw,
                    "Found": "-",
                    "Match": "-",
                    "Pages": "-",
                    "Font Size": "-",
                    "Note": "Manual check required",
                    "Verification": "Manual",
                    "Remark": remark_text,
                    "Remark URL": remark_link,
                    "Package Panel": package_panel,
                    "Procedure": procedure,
                    "__Term_HTML__": row.get("__Term_HTML__", ""),
                    "Image_Groups_Resolved": row.get("Image_Groups_Resolved", row.get("Image_Groups", [])),
                })
            else:
                for term in term_lines:
                    grouped[(requirement, spec, "Manual")].append({
                    "Term": term,
                    "Found": "-",
                    "Match": "-",
                    "Pages": "-",
                    "Font Size": "-",
                    "Note": "Manual check required",
                    "Verification": "Manual",
                    "Remark": remark_text,
                    "Remark URL": remark_link,
                    "Package Panel": package_panel,
                    "Procedure": procedure,
                    "__Term_HTML__": row.get("__Term_HTML__", ""),
                    "Image_Groups_Resolved": row.get("Image_Groups_Resolved", row.get("Image_Groups", [])),
                })
            continue

        STOPWORDS = {"in","en","na","de","la","el","em","da","do","di","du","of","and","y","et","the","a","an"}

        # Regular expression to extract word tokens 
        TOKEN_RE = re.compile(r"\b\w{2,}\b", re.UNICODE)

        # Auto-expand MADE IN into multilingual variants 
        MADE_IN_MAP = {
            "ENGLISH": ["made in"],
            "UK": ["made in"], "US": ["made in"],
            "SPANISH": ["hecho en", "fabricado en"], "LAAM SPANISH": ["hecho en", "fabricado en"],
            "CANADIAN FRENCH": ["fabriqué en", "fabriqué au"], "FRENCH": ["fabriqué en", "fabrique en"],
            "PORTUGUESE": ["feito em", "feito na", "fabricado em", "fabricado na"],
            "BRAZILIAN PORTUGUESE": ["feito no", "feito na", "feito em", "fabricado no", "fabricado na", "fabricado em"],
            "GERMAN": ["hergestellt in"],
            "ITALIAN": ["prodotto in", "fabbricato in"],
            "DUTCH": ["gemaakt in", "vervaardigd in"],
            "SWEDISH": ["tillverkad i"],
            "FINNISH": ["valmistettu"],
            "DANISH": ["fremstillet i", "produceret i"],
            "NORWEGIAN": ["produsert i", "fremstilt i"],
            "POLISH": ["wyprodukowano w"],
            "CZECH": ["vyrobeno v"],
            "SLOVAK": ["vyrobené v"],
            "HUNGARIAN": ["gyártva", "készült"],
            "RUSSIAN": ["сделано в", "произведено в"],
            "GREEK": ["κατασκευάζεται στην", "παράγεται στην"],
            "TURKISH": ["üretildiği", "üretim yeri", "ürün menşei"],
            "ARABIC": ["صنع في", "صناعة"]
        }

        def _is_bare_made_in(variant_norm: str) -> bool:
            toks = [t for t in TOKEN_RE.findall(variant_norm) if t not in STOPWORDS]
            if len(toks) <= 2:
                root = variant_norm
                return any(k in root for k in [
                    "made", "hecho", "fabrique", "fabriqu", "prodotto", "fabbricato",
                    "hergestellt", "gemaakt", "tillverkad", "valmistettu", "fremstillet",
                    "produceret", "produsert", "wyprodukowano", "vyroben", "gyártva", "készült",
                    "сделано", "произведено", "κατασκευ", "παράγ", "üret", "صنع", "صناعة"
                ])
            return False

        def _expand_made_in_variants(variants: list, lang_list: list) -> list:
            if not any(_is_bare_made_in(normalize_text(v)) for v in variants):
                return variants
            targets = [str(x).strip().upper() for x in (lang_list or []) if str(x).strip()]
            phrases = set()
            if targets:
                for lg in targets:
                    for p in MADE_IN_MAP.get(lg, []):
                        phrases.add(p)
            else:
                for lst in MADE_IN_MAP.values():
                    for p in lst:
                        phrases.add(p)
            out = list(dict.fromkeys(variants + list(phrases)))
            return out

        def _split_term_variants(term_raw: str):
            s = str(term_raw or "")
            s = s.replace("\r", "")
            s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)

            parts = []
            LANG_CUE = re.compile(
                (
                    r"\b("
                    r"made\s+in|"
                    r"hecho\s+en|"
                    r"fabricad\w*\s+en|"
                    r"fabriqu\w*\s+en|"
                    r"prodotto\s+in|"
                    r"fabbricato\s+in|"
                    r"hergestellt\s+in|"
                    r"feito\s+(?:em|no|na)|" 
                    r"gemaakt\s+in|"
                    r"tillverkad\s+i|"
                    r"valmistettu|"
                    r"fremstillet\s+i|"
                    r"produceret\s+i|"
                    r"produsert\s+i|"
                    r"vyroben[oaé]?|"
                    r"wyprodukowano\s+w|"
                    r"gyártva|"
                    r"készült|"
                    r"сделано\s+в|"
                    r"произведено\s+в|"
                    r"κατασκευ|"
                    r"παράγ|"
                    r"üret|"
                    r"صنع\s+في|"
                    r"صناعة"
                    r")\b"
                ),
                flags=re.I
            )

            primary = [p.strip() for p in re.split(r"[\n;|]+", s) if p.strip()]

            for chunk in primary:
                if re.search(r"\b(or|หรือ)\b", chunk, flags=re.I):
                    for seg in re.split(r"[\/,•·∙・／]+", chunk):
                        seg = seg.strip()
                        if seg:
                            parts.append(seg)
                    continue

                if LANG_CUE.search(chunk) and "." in chunk:
                    for seg in re.split(r"\.\s*", chunk):
                        seg = seg.strip(" ,/").strip()
                        if seg:
                            parts.append(seg)
                    continue

                for seg in re.split(r"[\/,•·∙・／\u2022\u00B7]+", chunk):
                    seg = seg.strip()
                    if seg:
                        parts.append(seg)

            return parts or [s.strip()]
        
        def _tokens_in_order(tokens, text_norm: str) -> bool:
            pos = 0
            for w in tokens:
                i = text_norm.find(w, pos)
                if i == -1:
                    return False
                pos = i + len(w)
            return True

        def _match_items_for_variant(variant_norm: str, all_texts, require_thailand: bool=False):
           
            STOPWORDS = {"in","en","na","de","la","el","em","da","do","di","du","of","and","y","et","the","a","an"}
            generic_drop = {"requirement", "address", "code"}
            keep = {"astm", "iso", "en71", "f963", "eu", "us", "uk", "br", "ca", "brazil", "canada", "canadian"}

            tokens = TOKEN_RE.findall(variant_norm)
            words = [w for w in tokens if (len(w) >= 3 or w in keep) and w not in generic_drop and w not in STOPWORDS]

            # ทางลัดสำหรับ “อายุ N+”
            m_age = re.fullmatch(r"\s*(\d{1,2})\s*[\+\＋]\s*", variant_norm)
            age_pat = None
            if m_age:
                n = m_age.group(1)
                age_pat = re.compile(rf"(?<!\w){re.escape(n)}\s*[\+\＋](?!\w)")
            
            risky = _is_risky_term(variant_norm)
            matched_items, pages = [], []
            pages_set = set()

            def _collapse_ws_hyphen(s: str) -> str:
                return re.sub(r"[\s\-]+", " ", s).strip()

            for text_norm, page_number, item in all_texts:
                src = (item.get("source") or "pdf").lower()
                hit = False

                if age_pat and age_pat.search(text_norm):
                    hit = True
                if variant_norm and variant_norm in text_norm:
                    hit = True
                elif words:
                    pos, ok = 0, True
                    for w in words:
                        i = text_norm.find(w, pos)
                        if i == -1:
                            ok = False; break
                        pos = i + len(w)
                    if ok:
                        hit = True

                elif risky:
                    allow_ocr_fuzzy = ( src == "ocr" and len(variant_norm) <= 6)
                    if src != "ocr" or allow_ocr_fuzzy:
                        if _fuzzy_ratio(_collapse_ws_hyphen(variant_norm), _collapse_ws_hyphen(text_norm)) >= 0.96:
                            hit = True

                if hit and require_thailand and not _must_contain_country_th(text_norm):
                    hit = False

                if hit:
                    matched_items.append(item)
                    pages_set.add(page_number)

            def _safe_sz(it):
                try: return float(it.get("size_mm") or 0.0)
                except Exception: 
                    return 0.0
                
            matched_items.sort(key=lambda it: (
                    str(it.get("level","")) == "line",
                    bool(it.get("bold")),
                    _safe_sz(it)
                ), reverse=True)
            
            # Page level fallback กรณีข้อความโดนตัดบรรทัดเลยไม่อยู่ใน item เดียว
            if len(words) >= 2:
                for pno, ptxt in page_norm_text.items():
                    if pno in pages_set:
                        continue
                    if _tokens_in_order(words, ptxt):
                        pages_set.add(pno)

            pages = sorted(pages_set)
            return matched_items, pages
        
        def _all_caps_items(items):
            got = False
            for it in items:
                t = (it.get("text") or "").strip()
                if t:
                    got = True
                    if not _is_all_caps_approx(t):
                        return False
            return got
        
        def _dedup_items(items):
            seen = set(); out = []
            for it in items:
                key = (normalize_text(it.get("text","")), (it.get("source") or "pdf").lower())
                if key in seen:
                    continue
                seen.add(key)
                out.append(it)
            return out

        # VERIFIED SECTION
        for term in term_lines:
            variants = _split_term_variants(term)
            variants = _expand_made_in_variants(variants, row.get("Language List", []))

            require_th = _extract_th_country_flag(term)
            
            union_pages = set()
            all_items= []
            best = {"items": [], "pages": [], "variant": ""}
            best_score = -1
            ul_keys    = ("underline", "ขีดเส้นใต้")
            no_ul_keys = ("no underline", "ไม่มีขีดเส้นใต้")

            for v in variants:
                v_norm = normalize_text(v)
                require_th = _extract_th_country_flag(term)
                items, pages = _match_items_for_variant(v_norm, all_texts, require_thailand=require_th)
                all_items.extend(items)
                union_pages.update(pages)
                score = len(items)

                if _contains_any(spec_lower, ul_keys):
                    score += 1000 * len([i for i in items if bool(i.get("underline"))])
                if _contains_any(spec_lower, no_ul_keys):
                    score += 1000 * len([i for i in items if not bool(i.get("underline"))])
                if score > best_score:
                    best_score = score
                    best = {"items": items, "pages": pages, "variant": v}

            # ใช้ best สำหรับตรวจรูปแบบ/ขนาดตัวอักษร
            matched_items = _dedup_items(all_items) or best["items"]

            # แต่ "การรายงานหน้า" ให้ใช้ union ของทุก variant
            found_pages_all = sorted(set(union_pages))

            found_flag   = "✅ Found" if found_pages_all else "❌ Not Found"
            match_result = "✔"
            notes        = []

            # ใส่เหตุผลเมื่อไม่พบ
            if not found_pages_all:
                notes.append("Not found on artwork pages")
                if bool(row.get("_HasImage", False)):
                    notes.append("Text may be image-only")

            # กรองตามสเปกจริง (underline/no-underline)
            if _contains_any(spec_lower, ul_keys):
                matched_items = [i for i in matched_items if bool(i.get("underline"))]
            if _contains_any(spec_lower, no_ul_keys):
                matched_items = [i for i in matched_items if not bool(i.get("underline"))]

            bolds       = [bool(i.get("bold", False)) for i in matched_items]
            underlines  = [bool(i.get("underline", False)) for i in matched_items]
            sizes_mm    = [_pick_size_mm(i) for i in matched_items]
            max_size_mm = max(sizes_mm) if sizes_mm else None
            
            # ---------- SALVAGE สำหรับสเปก Underline ----------
            underline_present = any(bool(i.get("underline", False)) for i in matched_items)

            if _contains_any(spec_lower, ("underline", "ขีดเส้นใต้")) and (not underline_present) and found_pages_all:
                under_tokens = set()
                try:
                    html_src = str(row.get("__Term_HTML__", "") or "")
                    under_frags = _extract_underlined_substrings(html_src)
                    for frag in under_frags:
                        v_norm = normalize_text(frag)
                        under_tokens.update(TOKEN_RE.findall(v_norm))
                except Exception:
                    pass

                # ถ้าไม่มี <u>…</u> ใน Excel → fallback เป็น token ของทุก variant
                if not under_tokens:
                    for v in variants:
                        v_norm = normalize_text(v)
                        under_tokens.update(TOKEN_RE.findall(v_norm))

                STOPWORDS = {"in","en","na","de","la","el","em","da","do","di","du","of","and","y","et","the","a","an"}
                under_tokens = {t for t in under_tokens if len(t) >= 2 and t not in STOPWORDS}

                # หาหลักฐานบนหน้าใดก็ได้ที่พบ requirement มี "คำใต้เส้น" ที่ตรงกับ under_tokens
                added = None
                added_page = None
                for text_norm, page_no, item in all_texts:
                    if page_no in found_pages_all and bool(item.get("underline")):
                        if (not under_tokens) or any(tok in text_norm for tok in under_tokens):
                            added = item
                            added_page = page_no
                            break
                if added is not None:
                    matched_items.append(added)
                    logger.info("    [SALVAGE] added underline evidence from page %s: %r (src=%s, size=%.2f mm)",
                        added_page,
                        (added.get("text") or "")[:80],
                        (added.get("source") or "pdf"),
                        _pick_size_mm(added))

            # คำนวณใหม่หลัง salvage
            bolds       = [bool(i.get("bold", False)) for i in matched_items]
            underlines  = [bool(i.get("underline", False)) for i in matched_items]
            sizes_mm    = [_pick_size_mm(i) for i in matched_items]
            max_size_mm = max(sizes_mm) if sizes_mm else None

            # ---- Bold ----
            bold_present = any(bolds)
            if _contains_any(spec_lower, ("bold", "ตัวหนา")) and not bold_present and found_pages_all:

                variant_norms = [normalize_text(vv) for vv in variants]
                variant_token_lists = [
                    [t for t in TOKEN_RE.findall(vn) if len(t) >= 2 and t not in STOPWORDS]
                    for vn in variant_norms
                ]

                added_bold = None
                added_bold_page = None

                for text_norm, page_no, item in all_texts:
                    if page_no not in found_pages_all:
                        continue

                    # bold จากเมทาดาต้า + ข้อความตรงอย่างน้อยหนึ่งภาษา
                    if bool(item.get("bold")):
                        if any( (vn and vn in text_norm) or (vtoks and all(tok in text_norm for tok in vtoks))
                                for vn, vtoks in zip(variant_norms, variant_token_lists) ):
                            added_bold = item
                            added_bold_page = page_no
                            break

                    # กู้แบบบรรทัด (line-level) ALL CAPS + ขนาดพอ และข้อความตรงอย่างน้อยหนึ่งภาษา
                    if str(item.get("level","")) == "line":
                        if any( (vn and vn in text_norm) or (vtoks and all(tok in text_norm for tok in vtoks))
                                for vn, vtoks in zip(variant_norms, variant_token_lists) ):
                            try:
                                if _is_all_caps_approx(item.get("text","")) and (_pick_size_mm(item) >= 1.2):
                                    added_bold = item
                                    added_bold_page = page_no
                                    break
                            except Exception:
                                pass

                if added_bold is not None:
                    matched_items.append(added_bold)
                    bold_present = True
                    notes.append("Bold evidence found on artwork page")

            # สรุปผล Bold 
            if _contains_any(spec_lower, ("bold", "ตัวหนา")) and not bold_present:
                match_result = "❌"
                notes.append("Not Bold")

            # ---- Underline ----
            underline_present = any(underlines)
            if _contains_any(spec_lower, ("no underline", "ไม่มีขีดเส้นใต้")):
                if underline_present:
                    match_result = "❌"; notes.append("Underline must be absent")
            elif _contains_any(spec_lower, ("underline", "ขีดเส้นใต้")):
                if not underline_present:
                    match_result = "❌"; notes.append("Underline Missing")

            # ---- All Caps ----
            if _contains_any(spec_lower, ("all caps", "ตัวพิมพ์ใหญ่ทั้งหมด", "ตัวใหญ่ทั้งหมด", "พิมพ์ใหญ่ทั้งหมด")):
                if not _all_caps_items(matched_items):
                    match_result = "❌"
                    notes.append("Not All Caps")

            # ---- Font size ----  
            thr_mm = _parse_threshold_to_mm(spec_lower)
            font_size_str = "-"
            size_note = None

            if thr_mm is not None:
                if matched_items and (max_size_mm is not None):
                    meas_str = _fmt_mm(max_size_mm)
                    spec_str = _fmt_mm(thr_mm)

                    size_note = f"size {meas_str} (spec ≥ {spec_str})"

                    if not _size_meets_threshold(max_size_mm, thr_mm):
                        match_result  = "❌"
                        font_size_str = "❌"
                    else:
                        font_size_str = "✔"
                else:
                    notes.append("No measurable text for font size")

            # ---- Pages ---- 
            page_str = _format_pages_for_output(found_pages_all)

            is_lang_row = (
                ("language code" in req_norm) or ("lang code" in req_norm) or ("ภาษา" in req_norm)
                or ("language code" in spec_norm) or ("lang code" in spec_norm) or ("ภาษา" in spec_norm)
            )

            expected_codes = set()
            for _src in (term_lines or []):
                expected_codes |= _detect_lang_codes_in_text(_src)
            expected_codes |= _detect_lang_codes_in_text(spec)

            detected_all = set()
            for _p, _codes in (page_langcodes or {}).items():
                detected_all |= (_codes or set())

            # ---- Notes ----
            is_found = bool(found_pages_all)
            is_pass  = (match_result == "✔")

            if is_found and is_pass:
                notes_to_show = [size_note] if size_note else []
            else:
                notes_to_show = _dedup_notes(notes + ([size_note] if size_note else []))

            note_str = ", ".join(notes_to_show) if notes_to_show else "-"

            # --- LOG หลักฐานการตรวจ (หลักฐานจาก PDF/OCR จริง) ---
            _log_evidence(
                term_text=term,
                spec_text=spec,
                found_pages_list=found_pages_all,
                matched_items=matched_items,
                thr_mm=thr_mm,
                max_size_mm=max_size_mm,
                page_str=page_str
            )

            grouped[(requirement, spec, "Verified")].append({
                "Term": term,
                "Found": found_flag,
                "Match": match_result if found_flag == "✅ Found" else "❌",
                "Pages": page_str if found_flag == "✅ Found" else "-",
                "Font Size": font_size_str if found_flag == "✅ Found" else "-",
                "Note": note_str,
                "Verification": "Verified",
                "Remark": remark_text,
                "Remark URL": remark_link,
                "Package Panel": package_panel,
                "Procedure": procedure,
                "__Term_HTML__": row.get("__Term_HTML__", ""),
                "Image_Groups_Resolved": row.get("Image_Groups_Resolved", row.get("Image_Groups", [])),
            })

    final_results = []
    for (requirement, spec, verification), items in grouped.items():
        for item in items:
            raw_term = item.get("Term", "")
            has_imgs = bool(item.get("Image_Groups_Resolved") or [])

            def _is_blank(s) -> bool:
                s = "" if s is None else str(s).strip()
                return s.lower() in ("", "nan", "none", "-")

            if _is_blank(raw_term) and has_imgs:
                term_display = ""
            elif _is_blank(raw_term):
                term_display = "-"
            else:
                term_display = str(raw_term)

            final_results.append({
                "Requirement": requirement,
                "Symbol/ Exact wording": term_display,
                "Specification": spec,
                "Package Panel": item.get("Package Panel", "-"),
                "Procedure": item.get("Procedure", "-"),
                "Remark": item.get("Remark", "-"),
                "Remark URL": item.get("Remark URL", "-"),
                "Found": item["Found"],
                "Match": item["Match"],
                "Pages": item["Pages"],
                "Font Size": item["Font Size"],
                "Note": item["Note"],
                "Verification": verification,
                "__Term_HTML__": item.get("__Term_HTML__", ""),
                "Image_Groups_Resolved": item.get("Image_Groups_Resolved", []),
            })

    return pd.DataFrame(final_results)